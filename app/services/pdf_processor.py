"""
PDF Processor Service
=====================
Handles PDF to Excel conversion with formatting preservation.

Features:
- Extract tables from PDF with structure preservation
- Extract formatting (colors, fonts) from PDF
- Merge multiple pages into single Excel sheet
- Convert to Excel with formatting applied

Author: datnguyentien@vietjetair.com
"""

from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import re
from datetime import datetime

import pandas as pd
import pdfplumber
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from app.core.logging import get_logger

logger = get_logger(__name__)


class PDFProcessor:
    """
    PDF processor for extracting tables and converting to Excel.
    
    Supports:
    - Table extraction with structure preservation
    - Formatting extraction (colors, fonts)
    - Multi-page merging into single sheet
    - Excel output with formatting applied
    """
    
    def __init__(self):
        """Initialize the PDF processor."""
        pass
    
    def extract_tables_with_style(
        self,
        pdf_path: Path | str
    ) -> Dict[str, Any]:
        """
        Extract tables from all pages with formatting information.
        
        Args:
            pdf_path: Path to PDF file.
            
        Returns:
            Dictionary with:
            - tables: List of tables (one per page) with cell bboxes
            - styles: Dictionary of cell styles by position
            - page_count: Number of pages
        """
        pdf_path = Path(pdf_path)
        
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        
        logger.info("Extracting tables from PDF", file=str(pdf_path))
        
        tables = []
        page_count = 0
        
        # Extract tables using pdfplumber with cell bboxes
        with pdfplumber.open(pdf_path) as pdf:
            page_count = len(pdf.pages)
            
            for page_num, page in enumerate(pdf.pages, 1):
                logger.info(f"Processing page {page_num}/{page_count}")
                
                # Try multiple extraction strategies for better results
                page_tables = []
                
                # Strategy 1: Try with lines_strict (best for tables with clear borders)
                try:
                    tables_strict = page.extract_tables(
                        table_settings={
                            "vertical_strategy": "lines_strict",
                            "horizontal_strategy": "lines_strict",
                            "snap_tolerance": 3,
                            "join_tolerance": 3,
                            "min_words_vertical": 1,
                            "min_words_horizontal": 1,
                            "intersection_tolerance": 3,
                        }
                    )
                    if tables_strict:
                        page_tables.extend(tables_strict)
                        logger.debug(f"Found {len(tables_strict)} tables with lines_strict on page {page_num}")
                except Exception as e:
                    logger.debug(f"lines_strict failed on page {page_num}: {e}")
                
                # Strategy 2: Try with text strategy (better for tables without clear borders)
                if not page_tables:
                    try:
                        tables_text = page.extract_tables(
                            table_settings={
                                "vertical_strategy": "text",
                                "horizontal_strategy": "text",
                                "snap_tolerance": 3,
                                "join_tolerance": 3,
                                "min_words_vertical": 1,
                                "min_words_horizontal": 1,
                            }
                        )
                        if tables_text:
                            page_tables.extend(tables_text)
                            logger.debug(f"Found {len(tables_text)} tables with text strategy on page {page_num}")
                    except Exception as e:
                        logger.debug(f"text strategy failed on page {page_num}: {e}")
                
                # Strategy 3: Try explicit lines if available
                if not page_tables:
                    try:
                        # Get vertical and horizontal lines from page
                        v_lines = page.vertical_edges
                        h_lines = page.horizontal_edges
                        
                        if v_lines and h_lines:
                            tables_explicit = page.extract_tables(
                                table_settings={
                                    "vertical_strategy": "explicit",
                                    "horizontal_strategy": "explicit",
                                    "explicit_vertical_lines": v_lines[:50] if len(v_lines) > 50 else v_lines,  # Limit to avoid performance issues
                                    "explicit_horizontal_lines": h_lines[:50] if len(h_lines) > 50 else h_lines,
                                    "snap_tolerance": 3,
                                    "join_tolerance": 3,
                                }
                            )
                            if tables_explicit:
                                page_tables.extend(tables_explicit)
                                logger.debug(f"Found {len(tables_explicit)} tables with explicit lines on page {page_num}")
                    except Exception as e:
                        logger.debug(f"explicit strategy failed on page {page_num}: {e}")
                
                if not page_tables:
                    logger.warning(f"No tables found on page {page_num} with any strategy")
                    continue
                
                # Remove duplicates and filter valid tables
                valid_tables = []
                seen_tables = set()
                
                for table in page_tables:
                    if not table or len(table) < 2:
                        continue
                    
                    # Create a hash of first few rows to detect duplicates
                    table_hash = tuple(tuple(row[:5]) for row in table[:3] if row)
                    if table_hash in seen_tables:
                        continue
                    seen_tables.add(table_hash)
                    
                    valid_tables.append(table)
                
                if not valid_tables:
                    logger.warning(f"No valid tables found on page {page_num}")
                    continue
                
                # Strategy: Find the main data table (largest with header indicators)
                main_table = None
                max_score = 0
                
                for table in valid_tables:
                    if not table or len(table) < 2:
                        continue
                    
                    # Calculate score: rows * cols
                    score = len(table) * (len(table[0]) if table[0] else 0)
                    
                    # Check all rows for header indicators (not just first row)
                    has_header = False
                    for row in table[:5]:  # Check first 5 rows
                        if not row:
                            continue
                        row_text = " ".join(str(cell) for cell in row if cell).upper()
                        if any(keyword in row_text for keyword in ["USER ID", "NAME", "WE", "TH", "FR", "SA", "SU", "MO", "TU"]):
                            has_header = True
                            score *= 1.5  # Bonus for likely header
                            break
                    
                    # Bonus for tables with many columns (date columns)
                    if len(table[0]) > 20:
                        score *= 1.2
                    
                    # Bonus for tables with many rows (data rows)
                    if len(table) > 10:
                        score *= 1.1
                    
                    if score > max_score:
                        max_score = score
                        main_table = table
                
                # Fallback to largest table if no good candidate
                if not main_table:
                    main_table = max(valid_tables, key=lambda t: len(t) * (len(t[0]) if t and t[0] else 0))
                
                if main_table:
                    # Extract cell bboxes for this table
                    table_bboxes = []
                    try:
                        # Get table boundaries from page
                        table_objs = page.find_tables()
                        if table_objs:
                            for table_obj in table_objs:
                                extracted = table_obj.extract()
                                if extracted and len(extracted) == len(main_table):
                                    # This is our table
                                    table_bboxes = table_obj.bbox
                                    break
                    except Exception as e:
                        logger.debug(f"Could not extract table bboxes: {e}")
                    
                    logger.info(
                        f"Selected table from page {page_num}",
                        rows=len(main_table),
                        cols=len(main_table[0]) if main_table else 0,
                        first_row_preview=str(main_table[0][:5]) if main_table else "N/A"
                    )
                    
                    tables.append({
                        "page": page_num,
                        "data": main_table,
                        "rows": len(main_table),
                        "cols": len(main_table[0]) if main_table else 0,
                        "bbox": table_bboxes
                    })
        
        # Extract formatting using PyMuPDF with table alignment
        try:
            styles = self._extract_styles(pdf_path, tables)
        except Exception as e:
            logger.warning(f"Could not extract styles: {e}", exc_info=True)
            styles = {}
        
        logger.info(
            "Tables extracted",
            pages=page_count,
            tables_found=len(tables),
            total_rows=sum(t["rows"] for t in tables),
            styles_found=len(styles)
        )
        
        return {
            "tables": tables,
            "styles": styles,
            "page_count": page_count
        }
    
    def _extract_styles(
        self, 
        pdf_path: Path, 
        tables: List[Dict[str, Any]]
    ) -> Dict[str, Dict[str, Any]]:
        """
        Extract cell formatting (colors, borders) from PDF.
        
        Args:
            pdf_path: Path to PDF file.
            tables: List of extracted tables with bbox info.
            
        Returns:
            Dictionary mapping cell position to style info.
            Format: {(page, row, col): {"fill_color": ..., "border_color": ...}}
        """
        styles = {}
        
        try:
            doc = fitz.open(pdf_path)
            
            # Build table structure map: page -> list of cell bboxes
            table_cells = {}  # {(page, row, col): bbox}
            
            # Extract cell positions from pdfplumber tables
            with pdfplumber.open(pdf_path) as pdf:
                for table_info in tables:
                    page_num = table_info["page"]
                    table_data = table_info["data"]
                    
                    if page_num > len(pdf.pages):
                        continue
                    
                    page_pdf = pdf.pages[page_num - 1]
                    
                    # Try to get cell bboxes
                    try:
                        table_objs = page_pdf.find_tables()
                        if table_objs:
                            for table_obj in table_objs:
                                extracted = table_obj.extract()
                                if len(extracted) == len(table_data):
                                    # Match cells
                                    for row_idx, row in enumerate(table_obj.rows):
                                        for col_idx, cell in enumerate(row.cells):
                                            if cell:
                                                table_cells[(page_num, row_idx, col_idx)] = cell
                                    break
                    except Exception as e:
                        logger.debug(f"Could not extract cell bboxes for page {page_num}: {e}")
            
            # Extract colors from PDF using multiple methods
            for page_num, page in enumerate(doc, 1):
                # Method 1: Get drawings (rectangles with colors)
                try:
                    drawings = page.get_drawings()
                    
                    for drawing in drawings:
                        # Check if this is a rectangle (cell background)
                        if "rect" in drawing:
                            rect = drawing["rect"]
                            fill_color = drawing.get("fill", None)
                            stroke_color = drawing.get("color", None)
                            
                            # Find which cell this rectangle belongs to
                            cell_pos = self._find_cell_for_rect(
                                rect, 
                                table_cells, 
                                page_num
                            )
                            
                            if cell_pos:
                                page_idx, row, col = cell_pos
                                key = f"{page_idx}_{row}_{col}"
                                
                                if key not in styles:
                                    styles[key] = {}
                                
                                # Convert fill color
                                if fill_color:
                                    if isinstance(fill_color, (list, tuple)) and len(fill_color) >= 3:
                                        r, g, b = int(fill_color[0] * 255), int(fill_color[1] * 255), int(fill_color[2] * 255)
                                        styles[key]["fill_color"] = f"{r:02x}{g:02x}{b:02x}"
                                    elif isinstance(fill_color, int):
                                        r = (fill_color >> 16) & 0xFF
                                        g = (fill_color >> 8) & 0xFF
                                        b = fill_color & 0xFF
                                        styles[key]["fill_color"] = f"{r:02x}{g:02x}{b:02x}"
                                
                                # Convert border/stroke color
                                if stroke_color:
                                    if isinstance(stroke_color, (list, tuple)) and len(stroke_color) >= 3:
                                        r, g, b = int(stroke_color[0] * 255), int(stroke_color[1] * 255), int(stroke_color[2] * 255)
                                        styles[key]["border_color"] = f"{r:02x}{g:02x}{b:02x}"
                                    elif isinstance(stroke_color, int):
                                        r = (stroke_color >> 16) & 0xFF
                                        g = (stroke_color >> 8) & 0xFF
                                        b = stroke_color & 0xFF
                                        styles[key]["border_color"] = f"{r:02x}{g:02x}{b:02x}"
                except Exception as e:
                    logger.debug(f"Error extracting drawings from page {page_num}: {e}")
                
                # Method 2: Extract colors from cell images (for cells we found)
                # This is a fallback if drawings don't capture all colors
                try:
                    for (page_idx, row, col), cell_bbox in table_cells.items():
                        if page_idx != page_num:
                            continue
                        
                        key = f"{page_idx}_{row}_{col}"
                        
                        # Skip if we already have color for this cell
                        if key in styles and "fill_color" in styles[key]:
                            continue
                        
                        # Get cell bbox
                        if hasattr(cell_bbox, 'bbox'):
                            cell_rect = cell_bbox.bbox
                        elif isinstance(cell_bbox, (list, tuple)) and len(cell_bbox) >= 4:
                            cell_rect = fitz.Rect(cell_bbox)
                        else:
                            continue
                        
                        # Render cell as image and detect dominant color
                        # Only do this for a sample of cells to avoid performance issues
                        if row < 50 and col < 30:  # Limit to first 50 rows, 30 cols
                            try:
                                # Render cell region
                                mat = fitz.Matrix(2, 2)  # 2x zoom for better color detection
                                pix = page.get_pixmap(matrix=mat, clip=cell_rect)
                                
                                # Get dominant color from image
                                if pix:
                                    # Sample center region of cell
                                    width, height = pix.width, pix.height
                                    center_x, center_y = width // 2, height // 2
                                    sample_size = min(10, width // 4, height // 4)
                                    
                                    # Get average color from center region
                                    r_sum, g_sum, b_sum = 0, 0, 0
                                    count = 0
                                    
                                    for y in range(max(0, center_y - sample_size), min(height, center_y + sample_size)):
                                        for x in range(max(0, center_x - sample_size), min(width, center_x + sample_size)):
                                            pixel = pix.pixel(x, y)
                                            if isinstance(pixel, int):
                                                r = (pixel >> 16) & 0xFF
                                                g = (pixel >> 8) & 0xFF
                                                b = pixel & 0xFF
                                            elif isinstance(pixel, (list, tuple)) and len(pixel) >= 3:
                                                r, g, b = pixel[0], pixel[1], pixel[2]
                                            else:
                                                continue
                                            
                                            r_sum += r
                                            g_sum += g
                                            b_sum += b
                                            count += 1
                                    
                                    if count > 0:
                                        r_avg = r_sum // count
                                        g_avg = g_sum // count
                                        b_avg = b_sum // count
                                        
                                        # Only store if color is significantly different from white
                                        if not (r_avg > 240 and g_avg > 240 and b_avg > 240):
                                            if key not in styles:
                                                styles[key] = {}
                                            styles[key]["fill_color"] = f"{r_avg:02x}{g_avg:02x}{b_avg:02x}"
                                    
                                    pix = None  # Free memory
                            except Exception as e:
                                logger.debug(f"Error extracting color from cell ({row}, {col}): {e}")
                except Exception as e:
                    logger.debug(f"Error in image-based color extraction for page {page_num}: {e}")
            
            doc.close()
            
        except Exception as e:
            logger.warning(f"Error extracting styles: {e}", exc_info=True)
        
        return styles
    
    def _find_cell_for_rect(
        self,
        rect: fitz.Rect,
        table_cells: Dict[Tuple[int, int, int], Any],
        page_num: int
    ) -> Optional[Tuple[int, int, int]]:
        """
        Find which cell a rectangle belongs to based on overlap.
        
        Args:
            rect: Rectangle bbox from PyMuPDF.
            table_cells: Dictionary of cell positions and bboxes.
            page_num: Page number.
            
        Returns:
            (page, row, col) tuple or None.
        """
        rect_center = fitz.Point(
            (rect.x0 + rect.x1) / 2,
            (rect.y0 + rect.y1) / 2
        )
        
        best_match = None
        best_overlap = 0
        
        for (page, row, col), cell_bbox in table_cells.items():
            if page != page_num:
                continue
            
            if hasattr(cell_bbox, 'bbox'):
                cell_rect = cell_bbox.bbox
            elif isinstance(cell_bbox, (list, tuple)) and len(cell_bbox) >= 4:
                cell_rect = fitz.Rect(cell_bbox)
            else:
                continue
            
            # Check if rect center is within cell
            if (cell_rect.x0 <= rect_center.x <= cell_rect.x1 and
                cell_rect.y0 <= rect_center.y <= cell_rect.y1):
                # Calculate overlap area
                overlap_x0 = max(rect.x0, cell_rect.x0)
                overlap_y0 = max(rect.y0, cell_rect.y0)
                overlap_x1 = min(rect.x1, cell_rect.x1)
                overlap_y1 = min(rect.y1, cell_rect.y1)
                
                if overlap_x1 > overlap_x0 and overlap_y1 > overlap_y0:
                    overlap_area = (overlap_x1 - overlap_x0) * (overlap_y1 - overlap_y0)
                    if overlap_area > best_overlap:
                        best_overlap = overlap_area
                        best_match = (page, row, col)
        
        return best_match
    
    def merge_pages_to_single_sheet(
        self,
        tables: List[Dict[str, Any]],
        remove_duplicate_headers: bool = True
    ) -> pd.DataFrame:
        """
        Merge tables from multiple pages into single DataFrame.
        
        Args:
            tables: List of table dictionaries from extract_tables_with_style.
            remove_duplicate_headers: If True, remove duplicate header rows.
            
        Returns:
            Combined DataFrame with all pages merged.
        """
        if not tables:
            return pd.DataFrame()
        
        logger.info(f"Merging {len(tables)} tables into single sheet")
        
        all_rows = []
        first_header = None
        header_row_idx = None
        
        for table_info in tables:
            table_data = table_info["data"]
            page_num = table_info.get("page", 0)
            
            if not table_data:
                logger.warning(f"Empty table data from page {page_num}")
                continue
            
            logger.info(
                f"Processing table from page {page_num}",
                rows=len(table_data),
                cols=len(table_data[0]) if table_data else 0
            )
            
            # Find header row (contains "User ID", "Name", or day names)
            # Don't assume first row is header - search for it
            current_header_row_idx = None
            current_header = None
            
            for idx, row in enumerate(table_data[:10]):  # Check first 10 rows
                if not row:
                    continue
                row_text = " ".join(str(cell) for cell in row if cell).upper()
                
                # Strong indicators: "USER ID" or "NAME"
                if "USER ID" in row_text or ("NAME" in row_text and len(row) > 5):
                    current_header_row_idx = idx
                    current_header = row
                    break
                
                # Weak indicators: day names (but not if it's just dates)
                if any(day in row_text for day in ["WE", "TH", "FR", "SA", "SU", "MO", "TU"]):
                    # Make sure it's not just a date row
                    if "USER" not in row_text and "NAME" not in row_text:
                        # This might be a date header row, check if previous row has "USER ID" or "NAME"
                        if idx > 0:
                            prev_row = table_data[idx - 1]
                            prev_row_text = " ".join(str(cell) for cell in prev_row if prev_row and cell).upper()
                            if "USER ID" in prev_row_text or "NAME" in prev_row_text:
                                current_header_row_idx = idx - 1
                                current_header = prev_row
                                break
                    else:
                        current_header_row_idx = idx
                        current_header = row
                        break
            
            # If this is the first table, save header
            if first_header is None:
                if current_header_row_idx is not None:
                    first_header = current_header
                    header_row_idx = current_header_row_idx
                    # Add all rows from this table
                    all_rows.extend(table_data)
                    logger.info(f"Found header at row {current_header_row_idx} on page {page_num}")
                else:
                    # No clear header found, use first row as header
                    first_header = table_data[0]
                    header_row_idx = 0
                    all_rows.extend(table_data)
                    logger.warning(f"No clear header found on page {page_num}, using first row")
            else:
                # Check if header matches
                header_matches = False
                if current_header_row_idx is not None and current_header:
                    # Compare headers (normalize for comparison)
                    first_header_norm = [str(c).strip().upper() for c in first_header if c]
                    current_header_norm = [str(c).strip().upper() for c in current_header if c]
                    
                    # Check if headers are similar (at least 80% match)
                    if len(first_header_norm) > 0 and len(current_header_norm) > 0:
                        min_len = min(len(first_header_norm), len(current_header_norm))
                        matches = sum(1 for i in range(min_len) if first_header_norm[i] == current_header_norm[i])
                        header_matches = matches / min_len >= 0.8
                
                if remove_duplicate_headers and header_matches:
                    # Skip duplicate header, add data rows only
                    # But keep "Work Group" rows and other non-header rows before header
                    if current_header_row_idx is not None:
                        # Add rows before header (like "Work Group" rows)
                        all_rows.extend(table_data[:current_header_row_idx])
                        # Skip header, add data rows
                        all_rows.extend(table_data[current_header_row_idx + 1:])
                        logger.info(f"Skipped duplicate header on page {page_num}, kept {current_header_row_idx} rows before header")
                    else:
                        # No header found, add all rows
                        all_rows.extend(table_data)
                else:
                    # Different header or keep all rows
                    all_rows.extend(table_data)
                    logger.info(f"Added all rows from page {page_num} (header doesn't match or keep all)")
        
        if not all_rows:
            return pd.DataFrame()
        
        # Normalize rows to have same column count
        max_cols = max(len(row) for row in all_rows) if all_rows else 0
        
        normalized_rows = []
        for row in all_rows:
            # Pad row to max_cols
            normalized_row = row + [""] * (max_cols - len(row))
            normalized_rows.append(normalized_row[:max_cols])
        
        # Create DataFrame
        # Try to find header row (contains "User ID", "Name", or day names)
        # Note: We already found header during merge, but we need to find it again in merged data
        header_row_idx = None
        
        if normalized_rows:
            logger.info(f"Creating DataFrame from {len(normalized_rows)} normalized rows")
            
            # Look for header row - prioritize rows with "User ID" or "Name"
            for idx, row in enumerate(normalized_rows[:15]):  # Check first 15 rows
                if not row:
                    continue
                row_text = " ".join(str(cell) for cell in row if cell).upper()
                
                # Strong indicators: "USER ID" or "NAME"
                if "USER ID" in row_text or ("NAME" in row_text and len(row) > 5):
                    header_row_idx = idx
                    logger.info(f"Found header row at index {idx}: {row[:5]}")
                    break
                
                # Weak indicators: day names (but not if it's just dates)
                if any(day in row_text for day in ["WE", "TH", "FR", "SA", "SU", "MO", "TU"]):
                    # Make sure it's not just a date row
                    if "USER" not in row_text and "NAME" not in row_text:
                        # This might be a date header row, check if previous row has "USER ID" or "NAME"
                        if idx > 0:
                            prev_row = normalized_rows[idx - 1]
                            if prev_row:
                                prev_row_text = " ".join(str(cell) for cell in prev_row if cell).upper()
                                if "USER ID" in prev_row_text or "NAME" in prev_row_text:
                                    header_row_idx = idx - 1
                                    logger.info(f"Found header row at index {idx-1} (before date row)")
                                    break
                    else:
                        header_row_idx = idx
                        logger.info(f"Found header row at index {idx} (contains day names)")
                        break
            
            if header_row_idx is not None:
                # Use found header row
                header_row = normalized_rows[header_row_idx]
                data_rows = normalized_rows[header_row_idx + 1:]
                
                logger.info(
                    f"Using header at row {header_row_idx}",
                    header_preview=str(header_row[:5]),
                    data_rows_count=len(data_rows)
                )
                
                # Clean header: remove None, use meaningful names for empty headers
                clean_header = []
                for i, h in enumerate(header_row):
                    if h and str(h).strip():
                        clean_header.append(str(h).strip())
                    else:
                        # Try to infer from position
                        if i == 0:
                            clean_header.append("User ID")
                        elif i == 1:
                            clean_header.append("Name")
                        else:
                            clean_header.append(f"Column_{i+1}")
                
                # Ensure all data rows have same length as header
                max_cols = len(clean_header)
                for i, row in enumerate(data_rows):
                    if len(row) < max_cols:
                        data_rows[i] = row + [""] * (max_cols - len(row))
                    elif len(row) > max_cols:
                        data_rows[i] = row[:max_cols]
                
                df = pd.DataFrame(data_rows, columns=clean_header)
            else:
                # No clear header found, check if first row looks like header
                first_row = normalized_rows[0]
                is_header = False
                
                if first_row and len(first_row) > 0:
                    # Check if first row looks like header (non-numeric, text-based)
                    non_numeric_count = sum(
                        1 for cell in first_row[:5]
                        if cell and isinstance(cell, str) and not cell.replace(".", "").replace("-", "").isdigit()
                    )
                    is_header = non_numeric_count >= 3  # At least 3 non-numeric cells in first 5
                
                if is_header:
                    logger.info("Using first row as header (looks like header)")
                    # Ensure all rows have same length
                    max_cols = len(first_row)
                    for i, row in enumerate(normalized_rows[1:]):
                        if len(row) < max_cols:
                            normalized_rows[i + 1] = row + [""] * (max_cols - len(row))
                        elif len(row) > max_cols:
                            normalized_rows[i + 1] = row[:max_cols]
                    
                    df = pd.DataFrame(normalized_rows[1:], columns=first_row)
                else:
                    # No header, use default column names
                    logger.warning("No header found, using default column names")
                    # Ensure all rows have same length
                    max_cols = max(len(row) for row in normalized_rows) if normalized_rows else 0
                    for i, row in enumerate(normalized_rows):
                        if len(row) < max_cols:
                            normalized_rows[i] = row + [""] * (max_cols - len(row))
                        elif len(row) > max_cols:
                            normalized_rows[i] = row[:max_cols]
                    
                    df = pd.DataFrame(normalized_rows)
                    df.columns = [f"Column_{i+1}" for i in range(len(df.columns))]
        else:
            logger.warning("No normalized rows to create DataFrame")
            df = pd.DataFrame()
        
        logger.info(
            "Pages merged",
            total_rows=len(df),
            total_cols=len(df.columns),
            df_empty=df.empty,
            sample_data=str(df.head(5).to_dict()) if len(df) > 0 else "No data"
        )
        
        # Validate DataFrame has data
        if df.empty:
            logger.error("Merged DataFrame is empty! No data extracted from PDF.")
            raise ValueError("No data extracted from PDF - DataFrame is empty after merging")
        
        if len(df) == 0:
            logger.error("Merged DataFrame has 0 rows! No data extracted from PDF.")
            raise ValueError("No data extracted from PDF - DataFrame has 0 rows after merging")
        
        # Log column names for debugging
        logger.info(f"DataFrame columns: {list(df.columns)}")
        
        # Log sample of first few rows
        if len(df) > 0:
            logger.info(f"First row sample: {df.iloc[0].to_dict() if len(df) > 0 else 'N/A'}")
            logger.info(f"Last row sample: {df.iloc[-1].to_dict() if len(df) > 0 else 'N/A'}")
        
        return df
    
    def convert_to_excel_with_formatting(
        self,
        df: pd.DataFrame,
        styles: Dict[str, Dict[str, Any]],
        output_path: Path | str,
        sheet_name: str = "Sheet1",
        tables: Optional[List[Dict[str, Any]]] = None,
        row_mapping: Optional[Dict[int, Tuple[int, int]]] = None
    ) -> Path:
        """
        Convert DataFrame to Excel with formatting applied.
        
        Args:
            df: DataFrame to convert.
            styles: Style dictionary from extract_tables_with_style.
            output_path: Output Excel file path.
            sheet_name: Name for the Excel sheet.
            tables: Original table data for style mapping.
            
        Returns:
            Path to created Excel file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        logger.info(
            "Converting to Excel with formatting",
            output=str(output_path),
            rows=len(df),
            cols=len(df.columns),
            styles_count=len(styles),
            df_preview=str(df.head(3).to_dict()) if len(df) > 0 else "Empty DataFrame"
        )
        
        # Validate DataFrame is not empty
        if df.empty:
            logger.error("DataFrame is empty! Cannot create Excel file.")
            raise ValueError("DataFrame is empty - no data to write to Excel")
        
        if len(df) == 0:
            logger.error("DataFrame has 0 rows! Cannot create Excel file.")
            raise ValueError("DataFrame has 0 rows - no data to write to Excel")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Count non-empty cells for logging
        non_empty_cells = 0
        total_cells = 0
        
        # Write data with formatting
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                total_cells += 1
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Count non-empty cells
                if value is not None and str(value).strip() != "":
                    non_empty_cells += 1
                
                # Apply default formatting
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Try to detect and format dates
                if value and isinstance(value, str):
                    # Try to parse date formats like "26", "27", "17.Dec.2025", etc.
                    date_value = self._parse_date(value)
                    if date_value:
                        cell.value = date_value
                        cell.number_format = "dd.mm.yyyy"
                
                # Apply header formatting for first row
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.border = Border(
                        left=Side(style="thin", color="000000"),
                        right=Side(style="thin", color="000000"),
                        top=Side(style="thin", color="000000"),
                        bottom=Side(style="thin", color="000000")
                    )
                else:
                    # Data row formatting - apply styles from PDF
                    # Map DataFrame row/col to style key
                    # Note: r_idx is 1-based (Excel row), r_idx-1 is DataFrame row (0-based)
                    df_row = r_idx - 1  # DataFrame row index (0-based, excluding header)
                    df_col = c_idx - 1  # DataFrame column index (0-based)
                    
                    # Get style key using row mapping
                    style_key = self._get_style_key(
                        df_row, 
                        df_col, 
                        tables,
                        row_mapping
                    )
                    
                    if style_key and style_key in styles:
                        style_info = styles[style_key]
                        
                        # Apply fill color
                        if "fill_color" in style_info:
                            fill_color = style_info["fill_color"]
                            # Convert hex to RGB for openpyxl
                            if len(fill_color) == 6:
                                cell.fill = PatternFill(
                                    start_color=fill_color,
                                    end_color=fill_color,
                                    fill_type="solid"
                                )
                        
                        # Apply border color
                        border_color = style_info.get("border_color", "000000")
                        fill_color_val = style_info.get("fill_color", "")
                        
                        # Check if this is a red border or red fill (common in roster)
                        is_red = False
                        if border_color and len(border_color) == 6:
                            is_red = border_color.lower() in ["ff0000", "dc143c", "c00000"]
                        if not is_red and fill_color_val and len(fill_color_val) == 6:
                            is_red = fill_color_val.lower() in ["ff0000", "dc143c", "c00000"]
                        
                        if is_red:
                            # Red border - make it more visible
                            border_side = Side(style="medium", color="FF0000")
                        elif len(border_color) == 6:
                            border_side = Side(style="thin", color=border_color)
                        else:
                            border_side = Side(style="thin", color="000000")
                        
                        cell.border = Border(
                            left=border_side,
                            right=border_side,
                            top=border_side,
                            bottom=border_side
                        )
                    else:
                        # Default border
                        cell.border = Border(
                            left=Side(style="thin", color="000000"),
                            right=Side(style="thin", color="000000"),
                            top=Side(style="thin", color="000000"),
                            bottom=Side(style="thin", color="000000")
                        )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save workbook
        wb.save(output_path)
        wb.close()
        
        # Verify file was created and has content
        if not output_path.exists():
            raise FileNotFoundError(f"Excel file was not created: {output_path}")
        
        file_size = output_path.stat().st_size
        if file_size < 1000:  # Less than 1KB is suspicious
            logger.warning(f"Excel file is very small ({file_size} bytes), may be empty or corrupted")
        
        logger.info(
            "Excel file created",
            file=str(output_path),
            file_size=file_size,
            total_cells=total_cells,
            non_empty_cells=non_empty_cells,
            empty_cells=total_cells - non_empty_cells
        )
        
        return output_path
    
    def _parse_date(self, value: str) -> Optional[datetime]:
        """
        Try to parse date from string.
        
        Args:
            value: String value that might be a date.
            
        Returns:
            datetime object or None.
        """
        if not value or not isinstance(value, str):
            return None
        
        value = value.strip()
        
        # Try common date formats
        date_formats = [
            "%d.%m.%Y",      # 17.12.2025
            "%d.%b.%Y",      # 17.Dec.2025
            "%d %b %Y",      # 17 Dec 2025
            "%d/%m/%Y",      # 17/12/2025
            "%Y-%m-%d",      # 2025-12-17
            "%d-%m-%Y",      # 17-12-2025
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
        
        # Try to parse day number only (assume current month/year context)
        if value.isdigit() and 1 <= int(value) <= 31:
            # This is just a day number, not a full date
            # Return None to keep as string
            return None
        
        return None
    
    def _get_style_key(
        self,
        row_idx: int,
        col_idx: int,
        tables: Optional[List[Dict[str, Any]]],
        row_mapping: Optional[Dict[int, Tuple[int, int]]] = None
    ) -> Optional[str]:
        """
        Get style key for a cell position.
        
        Args:
            row_idx: DataFrame row index (0-based, excluding header).
            col_idx: DataFrame column index (0-based).
            tables: Original table data.
            row_mapping: Mapping from DataFrame row to (page, original_row).
            
        Returns:
            Style key string or None.
        """
        if not tables:
            return None
        
        # Use row mapping if available
        if row_mapping and row_idx in row_mapping:
            page_num, orig_row = row_mapping[row_idx]
            # Format: "page_row_col"
            return f"{page_num}_{orig_row}_{col_idx}"
        
        # Fallback: assume first page
        return f"1_{row_idx}_{col_idx}"
    
    def convert_pdf_to_excel(
        self,
        pdf_path: Path | str,
        excel_path: Path | str,
        sheet_name: str = "Sheet1",
        merge_pages: bool = True
    ) -> Dict[str, Any]:
        """
        Complete workflow: Extract PDF tables and convert to Excel.
        
        Args:
            pdf_path: Path to input PDF file.
            excel_path: Path to output Excel file.
            sheet_name: Name for Excel sheet.
            merge_pages: If True, merge all pages into single sheet.
            
        Returns:
            Dictionary with conversion statistics.
        """
        pdf_path = Path(pdf_path)
        excel_path = Path(excel_path)
        
        logger.info("Starting PDF to Excel conversion", pdf=str(pdf_path))
        
        # Extract tables and styles
        extraction_result = self.extract_tables_with_style(pdf_path)
        tables = extraction_result["tables"]
        styles = extraction_result["styles"]
        
        if not tables:
            raise ValueError("No tables found in PDF")
        
        # Build row mapping for style application
        # Map: DataFrame row index -> (page, original_row_in_page)
        row_mapping = {}
        current_df_row = 0
        
        if merge_pages:
            # Build mapping while merging
            first_header = None
            for table_info in tables:
                table_data = table_info["data"]
                page_num = table_info["page"]
                
                if not table_data:
                    continue
                
                current_header = table_data[0] if table_data else None
                
                if first_header is None:
                    first_header = current_header
                    # First table: include header
                    for orig_row in range(len(table_data)):
                        row_mapping[current_df_row] = (page_num, orig_row)
                        current_df_row += 1
                else:
                    # Subsequent tables: skip duplicate header
                    if current_header == first_header:
                        for orig_row in range(1, len(table_data)):
                            row_mapping[current_df_row] = (page_num, orig_row)
                            current_df_row += 1
                    else:
                        for orig_row in range(len(table_data)):
                            row_mapping[current_df_row] = (page_num, orig_row)
                            current_df_row += 1
            
            df = self.merge_pages_to_single_sheet(tables)
        else:
            # Use first table only
            df = pd.DataFrame(tables[0]["data"][1:], columns=tables[0]["data"][0])
            # Build simple mapping
            for orig_row in range(1, len(tables[0]["data"])):
                row_mapping[orig_row - 1] = (tables[0]["page"], orig_row)
        
        # Convert to Excel with formatting
        self.convert_to_excel_with_formatting(
            df, 
            styles, 
            excel_path, 
            sheet_name, 
            tables,
            row_mapping
        )
        
        stats = {
            "pages_processed": extraction_result["page_count"],
            "tables_found": len(tables),
            "total_rows": len(df),
            "total_cols": len(df.columns),
            "excel_path": str(excel_path)
        }
        
        logger.info("PDF to Excel conversion completed", **stats)
        
        return stats
    
    def get_page_count(self, pdf_path: Path | str) -> int:
        """
        Get number of pages in PDF.
        
        Args:
            pdf_path: Path to PDF file.
            
        Returns:
            Number of pages.
        """
        pdf_path = Path(pdf_path)
        
        with pdfplumber.open(pdf_path) as pdf:
            return len(pdf.pages)
    
    def extract_text(self, pdf_path: Path | str, page_num: Optional[int] = None) -> str:
        """
        Extract text from PDF (for debugging/inspection).
        
        Args:
            pdf_path: Path to PDF file.
            page_num: Specific page number (None for all pages).
            
        Returns:
            Extracted text.
        """
        pdf_path = Path(pdf_path)
        text_parts = []
        
        with pdfplumber.open(pdf_path) as pdf:
            if page_num:
                pages = [pdf.pages[page_num - 1]]
            else:
                pages = pdf.pages
            
            for page in pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
        
        return "\n".join(text_parts)
