"""
Excel Processor Service
=======================
Handles reading and writing Excel files for roster processing.

Author: datnguyentien@vietjetair.com
"""

from typing import Dict, List, Any, Optional
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from app.core.logging import get_logger

logger = get_logger(__name__)


class ExcelProcessor:
    """
    Excel file processor for reading and writing roster data.
    
    Supports both .xlsx and .xls formats.
    
    Example:
        >>> processor = ExcelProcessor()
        >>> df = processor.read_workbook("roster.xlsx", "Sheet1")
        >>> sheets = processor.get_sheet_names("roster.xlsx")
    """
    
    def __init__(self):
        """Initialize the Excel processor."""
        pass
    
    def get_sheet_names(self, file_path: Path | str) -> List[str]:
        """
        Get all sheet names from an Excel workbook.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            List of sheet names.
            
        Raises:
            FileNotFoundError: If file doesn't exist.
            ValueError: If file format is invalid.
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        ext = file_path.suffix.lower()
        
        if ext == ".xlsx":
            wb = load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        elif ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(file_path)
            return wb.sheet_names()
        else:
            raise ValueError(f"Unsupported file format: {ext}")
    
    def read_workbook(
        self,
        file_path: Path | str,
        sheet_name: str,
        header_row: int = 0,
        skip_rows: Optional[List[int]] = None
    ) -> pd.DataFrame:
        """
        Read a sheet from an Excel workbook into a DataFrame.
        
        Args:
            file_path: Path to the Excel file.
            sheet_name: Name of the sheet to read.
            header_row: Row index containing headers (0-based).
            skip_rows: Optional list of row indices to skip.
            
        Returns:
            pandas DataFrame with sheet data.
            
        Raises:
            FileNotFoundError: If file doesn't exist.
            ValueError: If sheet doesn't exist or format is invalid.
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        logger.info(
            "Reading workbook",
            file=str(file_path),
            sheet=sheet_name
        )
        
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header_row,
                skiprows=skip_rows
            )
            
            logger.info(
                "Workbook read successfully",
                rows=len(df),
                columns=len(df.columns)
            )
            
            return df
            
        except ValueError as e:
            if "Worksheet" in str(e):
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            raise
    
    def write_workbook(
        self,
        df: pd.DataFrame,
        file_path: Path | str,
        sheet_name: str = "Sheet1",
        index: bool = False
    ) -> Path:
        """
        Write a DataFrame to an Excel workbook.
        
        Args:
            df: DataFrame to write.
            file_path: Output file path.
            sheet_name: Name for the sheet.
            index: Whether to include DataFrame index.
            
        Returns:
            Path to the written file.
        """
        file_path = Path(file_path)
        
        # Ensure parent directory exists
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        logger.info(
            "Writing workbook",
            file=str(file_path),
            sheet=sheet_name,
            rows=len(df)
        )
        
        df.to_excel(
            file_path,
            sheet_name=sheet_name,
            index=index,
            engine="openpyxl"
        )
        
        logger.info("Workbook written successfully", file=str(file_path))
        
        return file_path
    
    def preview_sheet(
        self,
        file_path: Path | str,
        sheet_name: str,
        max_rows: int = 10,
        max_cols: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Preview a sheet with limited rows for inspection.
        
        Args:
            file_path: Path to the Excel file.
            sheet_name: Name of the sheet to preview.
            max_rows: Maximum number of data rows to return.
            max_cols: Maximum number of columns (None for all).
            
        Returns:
            Dictionary with headers, rows, and metadata.
        """
        df = self.read_workbook(file_path, sheet_name)
        
        total_rows = len(df)
        total_cols = len(df.columns)
        
        # Limit columns if specified
        if max_cols and max_cols < total_cols:
            df = df.iloc[:, :max_cols]
        
        # Get headers
        headers = [str(col) for col in df.columns]
        
        # Get preview rows (convert to strings for JSON serialization)
        preview_df = df.head(max_rows)
        rows = [
            [str(val) if pd.notna(val) else "" for val in row]
            for row in preview_df.values.tolist()
        ]
        
        return {
            "headers": headers,
            "rows": rows,
            "total_rows": total_rows,
            "total_cols": total_cols,
            "preview_rows": len(rows)
        }
    
    def merge_sheets(
        self,
        file_path: Path | str,
        sheet_names: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """
        Merge multiple sheets into a single DataFrame.
        
        Args:
            file_path: Path to the Excel file.
            sheet_names: List of sheets to merge (None for all).
            
        Returns:
            Combined DataFrame with all sheets.
        """
        file_path = Path(file_path)
        
        if sheet_names is None:
            sheet_names = self.get_sheet_names(file_path)
        
        dfs = []
        for sheet in sheet_names:
            df = self.read_workbook(file_path, sheet)
            df["_source_sheet"] = sheet
            dfs.append(df)
        
        if not dfs:
            return pd.DataFrame()
        
        return pd.concat(dfs, ignore_index=True)
    
    def copy_with_mapping(
        self,
        source_path: Path | str,
        dest_path: Path | str,
        sheet_name: str,
        mapped_df: pd.DataFrame
    ) -> Path:
        """
        Copy an Excel file and replace a sheet with mapped data.
        
        Preserves formatting from the original file where possible.
        
        Args:
            source_path: Original Excel file.
            dest_path: Destination file path.
            sheet_name: Sheet to replace.
            mapped_df: DataFrame with mapped data.
            
        Returns:
            Path to the new file.
        """
        source_path = Path(source_path)
        dest_path = Path(dest_path)
        
        # Load original workbook
        wb = load_workbook(source_path)
        
        # Get or create sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Clear existing data
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(sheet_name)
        
        # Write new data
        for r_idx, row in enumerate(dataframe_to_rows(mapped_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Save to destination
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(dest_path)
        wb.close()
        
        logger.info(
            "Workbook copied with mapping",
            source=str(source_path),
            dest=str(dest_path)
        )
        
        return dest_path
    
    def validate_structure(
        self,
        file_path: Path | str,
        required_columns: List[str],
        sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Validate that a workbook has the required structure.
        
        Args:
            file_path: Path to the Excel file.
            required_columns: List of required column names.
            sheet_name: Specific sheet to check (first sheet if None).
            
        Returns:
            Validation result with found/missing columns.
        """
        file_path = Path(file_path)
        
        sheets = self.get_sheet_names(file_path)
        
        if sheet_name is None:
            sheet_name = sheets[0]
        elif sheet_name not in sheets:
            return {
                "valid": False,
                "error": f"Sheet '{sheet_name}' not found",
                "available_sheets": sheets
            }
        
        df = self.read_workbook(file_path, sheet_name)
        actual_columns = [str(col).lower() for col in df.columns]
        
        found = []
        missing = []
        
        for req_col in required_columns:
            if req_col.lower() in actual_columns:
                found.append(req_col)
            else:
                missing.append(req_col)
        
        return {
            "valid": len(missing) == 0,
            "sheet": sheet_name,
            "found_columns": found,
            "missing_columns": missing,
            "all_columns": list(df.columns)
        }
    
    def write_multi_sheet_workbook(
        self,
        sheets_data: Dict[str, pd.DataFrame],
        file_path: Path | str,
        index: bool = False
    ) -> Path:
        """
        Write multiple DataFrames to separate sheets in a single workbook.
        
        Args:
            sheets_data: Dictionary mapping sheet names to DataFrames.
            file_path: Output file path.
            index: Whether to include DataFrame index.
            
        Returns:
            Path to the written file.
            
        Example:
            >>> processor.write_multi_sheet_workbook({
            ...     "Sheet1": df1,
            ...     "Sheet2": df2,
            ... }, "output.xlsx")
        """
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        logger.info(
            "Writing multi-sheet workbook",
            file=str(file_path),
            sheets=list(sheets_data.keys())
        )
        
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=index)
                logger.info(f"Written sheet: {sheet_name}", rows=len(df))
        
        logger.info("Multi-sheet workbook written successfully", file=str(file_path))
        return file_path
    
    def read_all_sheets(
        self,
        file_path: Path | str,
        sheet_names: Optional[List[str]] = None
    ) -> Dict[str, pd.DataFrame]:
        """
        Read multiple sheets from a workbook.
        
        Args:
            file_path: Path to the Excel file.
            sheet_names: List of sheets to read (None for all).
            
        Returns:
            Dictionary mapping sheet names to DataFrames.
        """
        file_path = Path(file_path)
        
        if sheet_names is None:
            sheet_names = self.get_sheet_names(file_path)
        
        result = {}
        for sheet in sheet_names:
            try:
                result[sheet] = self.read_workbook(file_path, sheet)
                logger.info(f"Read sheet: {sheet}", rows=len(result[sheet]))
            except Exception as e:
                logger.error(f"Error reading sheet {sheet}: {e}")
                continue
        
        return result
    
    def copy_workbook_with_mappings(
        self,
        source_path: Path | str,
        dest_path: Path | str,
        mapped_sheets: Dict[str, pd.DataFrame],
        preserve_unmapped: bool = True
    ) -> Path:
        """
        Copy workbook and replace multiple sheets with mapped data.
        
        Args:
            source_path: Original Excel file.
            dest_path: Destination file path.
            mapped_sheets: Dict of sheet_name -> mapped DataFrame.
            preserve_unmapped: Keep sheets that aren't in mapped_sheets.
            
        Returns:
            Path to the new file.
        """
        source_path = Path(source_path)
        dest_path = Path(dest_path)
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Load original workbook
        wb = load_workbook(source_path)
        
        for sheet_name, mapped_df in mapped_sheets.items():
            # Get or create sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Clear existing data
                for row in ws.iter_rows():
                    for cell in row:
                        cell.value = None
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Write new data
            for r_idx, row in enumerate(dataframe_to_rows(mapped_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Remove unmapped sheets if not preserving
        if not preserve_unmapped:
            for sheet_name in wb.sheetnames:
                if sheet_name not in mapped_sheets:
                    del wb[sheet_name]
        
        wb.save(dest_path)
        wb.close()
        
        logger.info(
            "Workbook copied with multi-sheet mappings",
            source=str(source_path),
            dest=str(dest_path),
            sheets=list(mapped_sheets.keys())
        )
        
        return dest_path

