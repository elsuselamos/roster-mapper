"""
XLS to XLSX Converter
=====================
Convert legacy .xls files to .xlsx format using LibreOffice.

Required: LibreOffice must be installed (soffice command available).
On Cloud Run, install via Dockerfile: apt-get install libreoffice-calc

Author: Vietjet AMO - IT Department
"""

import os
import subprocess
import shutil
from pathlib import Path
from typing import Optional, Tuple

from app.core.logging import get_logger

logger = get_logger(__name__)


def is_xls_file(filename: str) -> bool:
    """
    Check if file is a legacy .xls format.
    
    Args:
        filename: Filename to check
        
    Returns:
        True if .xls file, False otherwise
    """
    return filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx")


def check_libreoffice_available() -> bool:
    """
    Check if LibreOffice is available on the system.
    
    Returns:
        True if available, False otherwise
    """
    try:
        result = subprocess.run(
            ["soffice", "--version"],
            capture_output=True,
            text=True,
            timeout=10
        )
        return result.returncode == 0
    except (subprocess.SubprocessError, FileNotFoundError):
        return False


def convert_xls_to_xlsx(
    xls_path: str | Path,
    output_dir: Optional[str | Path] = None,
    cleanup_original: bool = False
) -> Tuple[Path, bool]:
    """
    Convert .xls file to .xlsx using LibreOffice.
    
    Args:
        xls_path: Path to .xls file
        output_dir: Directory for output (defaults to same directory as input)
        cleanup_original: Whether to delete the original .xls file after conversion
        
    Returns:
        Tuple of (output_path, success)
        
    Raises:
        FileNotFoundError: If input file doesn't exist
        RuntimeError: If LibreOffice is not available or conversion fails
    """
    xls_path = Path(xls_path)
    
    if not xls_path.exists():
        raise FileNotFoundError(f"Input file not found: {xls_path}")
    
    if not is_xls_file(str(xls_path)):
        logger.warning(
            "File is not .xls format, returning as-is",
            file=str(xls_path)
        )
        return xls_path, True
    
    # Determine output directory
    out_dir = Path(output_dir) if output_dir else xls_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    
    # Expected output path
    xlsx_path = out_dir / (xls_path.stem + ".xlsx")
    
    logger.info(
        "Converting XLS to XLSX",
        input=str(xls_path),
        output_dir=str(out_dir)
    )
    
    try:
        # Run LibreOffice conversion
        cmd = [
            "soffice",
            "--headless",
            "--convert-to", "xlsx",
            "--outdir", str(out_dir),
            str(xls_path)
        ]
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=120  # 2 minute timeout
        )
        
        if result.returncode != 0:
            logger.error(
                "LibreOffice conversion failed",
                returncode=result.returncode,
                stderr=result.stderr,
                stdout=result.stdout
            )
            raise RuntimeError(f"LibreOffice conversion failed: {result.stderr}")
        
        # Verify output file exists
        if not xlsx_path.exists():
            # Sometimes LibreOffice outputs to a slightly different name
            possible_outputs = list(out_dir.glob(f"{xls_path.stem}*.xlsx"))
            if possible_outputs:
                xlsx_path = possible_outputs[0]
            else:
                raise RuntimeError(
                    f"Conversion completed but output file not found: {xlsx_path}"
                )
        
        logger.info(
            "XLS to XLSX conversion successful",
            input=str(xls_path),
            output=str(xlsx_path),
            output_size=xlsx_path.stat().st_size
        )
        
        # Cleanup original if requested
        if cleanup_original and xls_path.exists():
            xls_path.unlink()
            logger.info("Original XLS file deleted", file=str(xls_path))
        
        return xlsx_path, True
        
    except subprocess.TimeoutExpired:
        logger.error(
            "LibreOffice conversion timeout",
            input=str(xls_path),
            timeout_seconds=120
        )
        raise RuntimeError("LibreOffice conversion timed out")
        
    except FileNotFoundError:
        logger.error(
            "LibreOffice not found - install libreoffice-calc",
            command="soffice"
        )
        raise RuntimeError(
            "LibreOffice not installed. Install with: apt-get install libreoffice-calc"
        )


def convert_xls_to_xlsx_fallback(
    xls_path: str | Path,
    output_dir: Optional[str | Path] = None
) -> Tuple[Path, bool]:
    """
    Try to convert .xls to .xlsx with fallback to xlrd+openpyxl.
    
    This is a fallback method if LibreOffice is not available.
    Note: This may lose some formatting compared to LibreOffice.
    
    Args:
        xls_path: Path to .xls file
        output_dir: Directory for output
        
    Returns:
        Tuple of (output_path, success)
    """
    xls_path = Path(xls_path)
    out_dir = Path(output_dir) if output_dir else xls_path.parent
    xlsx_path = out_dir / (xls_path.stem + ".xlsx")
    
    try:
        import xlrd
        from openpyxl import Workbook
        
        # Read .xls file
        xls_wb = xlrd.open_workbook(str(xls_path))
        
        # Create new .xlsx workbook
        xlsx_wb = Workbook()
        xlsx_wb.remove(xlsx_wb.active)  # Remove default sheet
        
        for sheet_name in xls_wb.sheet_names():
            xls_sheet = xls_wb.sheet_by_name(sheet_name)
            xlsx_sheet = xlsx_wb.create_sheet(title=sheet_name)
            
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
        
        xlsx_wb.save(str(xlsx_path))
        
        logger.info(
            "XLS to XLSX conversion (fallback) successful",
            input=str(xls_path),
            output=str(xlsx_path)
        )
        
        return xlsx_path, True
        
    except ImportError as e:
        logger.error("xlrd not installed for fallback conversion", error=str(e))
        raise RuntimeError("xlrd not installed. Install with: pip install xlrd")
        
    except Exception as e:
        logger.error("Fallback conversion failed", error=str(e))
        raise RuntimeError(f"XLS conversion failed: {e}")


def smart_convert_xls(
    xls_path: str | Path,
    output_dir: Optional[str | Path] = None,
    prefer_libreoffice: bool = True
) -> Tuple[Path, bool]:
    """
    Smart conversion that tries LibreOffice first, then falls back to xlrd.
    
    Args:
        xls_path: Path to .xls file
        output_dir: Directory for output
        prefer_libreoffice: Try LibreOffice first (better quality)
        
    Returns:
        Tuple of (output_path, success)
    """
    xls_path = Path(xls_path)
    
    if not is_xls_file(str(xls_path)):
        return xls_path, True
    
    if prefer_libreoffice and check_libreoffice_available():
        try:
            return convert_xls_to_xlsx(xls_path, output_dir)
        except RuntimeError:
            logger.warning("LibreOffice failed, trying fallback conversion")
    
    # Try fallback
    return convert_xls_to_xlsx_fallback(xls_path, output_dir)


